"""Audit the public BAFU Excel inventory without importing it into Brightway.

The output intentionally contains only aggregate counts and matching activity
metadata; it does not reproduce inventory exchanges.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def as_number(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def audit_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(f"Expected one worksheet; found {workbook.sheetnames}")

    worksheet = workbook[workbook.sheetnames[0]]
    database_name = None
    current_activity: dict[str, Any] | None = None
    exchange_headers: list[str | None] | None = None

    activity_count = 0
    exchange_types: Counter[str] = Counter()
    uncertainty_types: Counter[str] = Counter()
    uncertain_by_exchange_type: Counter[str] = Counter()
    swiss_low_voltage_candidates: list[dict[str, Any]] = []

    def finish_activity() -> None:
        if not current_activity:
            return
        name = str(current_activity.get("name") or "")
        product = str(current_activity.get("reference product") or "")
        location = str(current_activity.get("location") or "")
        if (
            location == "CH"
            and "electricity" in name.lower()
            and "low voltage" in name.lower()
        ):
            swiss_low_voltage_candidates.append(
                {
                    "name": name,
                    "reference_product": product,
                    "location": location,
                    "unit": current_activity.get("unit"),
                    "code": current_activity.get("code"),
                }
            )

    for row in worksheet.iter_rows(values_only=True):
        first = row[0]
        second = row[1] if len(row) > 1 else None

        if first == "Database":
            database_name = second
            continue

        if first == "Activity":
            finish_activity()
            current_activity = {"name": second}
            activity_count += 1
            exchange_headers = None
            continue

        if current_activity is None:
            continue

        if first in {"code", "location", "reference product", "unit"}:
            current_activity[str(first)] = second
            continue

        if first == "name" and "amount" in row and "type" in row:
            exchange_headers = [
                str(value) if value is not None else None for value in row
            ]
            continue

        if exchange_headers is None or first is None:
            continue

        exchange = {
            header: value
            for header, value in zip(exchange_headers, row)
            if header is not None
        }
        exchange_type = str(exchange.get("type") or "missing")
        exchange_types[exchange_type] += 1

        uncertainty = exchange.get("uncertainty type")
        uncertainty_number = as_number(uncertainty)
        uncertainty_key = (
            str(int(uncertainty_number))
            if uncertainty_number is not None and uncertainty_number.is_integer()
            else str(uncertainty or "missing")
        )
        uncertainty_types[uncertainty_key] += 1
        if uncertainty_number is not None and uncertainty_number > 0:
            uncertain_by_exchange_type[exchange_type] += 1

    finish_activity()
    workbook.close()

    return {
        "workbook": str(path),
        "file_size_bytes": path.stat().st_size,
        "sheet": worksheet.title,
        "rows": worksheet.max_row,
        "columns": worksheet.max_column,
        "database_name": database_name,
        "activities": activity_count,
        "exchange_types": dict(exchange_types),
        "uncertainty_types": dict(uncertainty_types),
        "uncertain_by_exchange_type": dict(uncertain_by_exchange_type),
        "swiss_low_voltage_candidates": swiss_low_voltage_candidates,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    result = audit_workbook(args.workbook)
    rendered = json.dumps(result, indent=2, sort_keys=True)
    print(rendered)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
